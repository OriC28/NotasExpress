from openpyxl import load_workbook
from threading import Thread

from PyQt6.QtWidgets import QWidget, QApplication, QPushButton, QMessageBox, QFileDialog
from PyQt6 import uic, QtGui

from Lib.Extraction import Extraction

class Generator:
	def __init__(self):
		self.file_path = None
		self.extr = None
		self.workbook = None
		self.message = QMessageBox 
		self.program = uic.loadUi("GUI/GUI.ui")
		self.init_gui()

	def init_gui(self):
		self.program.AddButton.clicked.connect(self.select_excel_file)
		self.program.CbYearSection.currentIndexChanged.connect(self.get_notes)
		self.program.show()

	def select_excel_file(self):
		try:
			file_path, filter = QFileDialog.getOpenFileName(self.program, 'Open file', '', 'Excel files (*.xlsx)')
			if file_path:
				self.file_path = file_path

				self.workbook = load_workbook(file_path, data_only=True) # CARGANDO ARCHIVO EXCEL

				self.program.CbYearSection.addItems(self.workbook.sheetnames)
			
				self.program.DirectoryEntry.setText(file_path)
				self.program.CbYearSection.setEnabled(True)

		except Exception as e:
			print(e)

	def get_notes(self):
		try:
			sheet_choiced = self.program.CbYearSection.currentIndex()

			self.extr = Extraction(self.workbook, sheet_choiced)

			first_table = self.extr.find_start_end_table(14, 30)

			second_table = self.extr.find_start_end_table(int(first_table[1])+1, 60)

			notes = self.extr.save_notes_subjects(14, first_table)

			notes.update(self.extr.save_notes_subjects(str(int(second_table[0])-2), second_table))

			print(f'\n{notes}\n')
		
		except Exception as e:
			print(e)

if __name__ == '__main__':
	app = QApplication([])
	generator = Generator()
	app.exec()