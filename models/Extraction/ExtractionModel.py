from openpyxl import load_workbook
import re

# NotasExpress
# Copyright (C) 2025 Escuela Técnica Industrial Nacional Capitán Giovanni Ferrareis 
# Licenciado bajo la GNU GPLv3. Ver <https://www.gnu.org/licenses/>.

class ExtractionModel:
    """
    Clase para extraer datos de una hoja de cálculo de Excel.

    Attributes:
        file_path (str): Ruta del archivo Excel.
        choiced (int): Índice de la hoja a utilizar. Por defecto es 0.
        workbook (Workbook): Objeto Workbook de openpyxl.
        sheets (list): Lista de nombres de las hojas en el archivo Excel.
        sheet_choiced (Worksheet): Hoja seleccionada para la extracción de datos.
    """

    def __init__(self, file_path: str, choiced=0):
        """
        Inicializa la clase ExtractionModel.

        Args:
            file_path (str): Ruta del archivo Excel.
            choiced (int, optional): Índice de la hoja a utilizar. Por defecto es 0.
        """
        self.file_path = file_path
        self.choiced = choiced
        self.workbook = load_workbook(self.file_path, data_only=True)
        self.sheets = self.workbook.sheetnames
        self.sheet_choiced = self.workbook[self.sheets[choiced]]
    
    def find_start(self, aprox_start: int, aprox_end: int):
        """
        Encuentra la fila de inicio de la tabla basándose en un patrón de texto.

        Args:
            aprox_start (int): Fila aproximada donde comienza la búsqueda.
            aprox_end (int): Fila aproximada donde termina la búsqueda.

        Returns:
            int or False: Fila de inicio si se encuentra, False si no se encuentra.
        """
        start = None
        for n in range(aprox_start, aprox_end):
            if re.findall(r'^V-|^CE-', str(self.sheet_choiced['C' + str(n)].value)):
                start = n
                break
        return start if start!=None else False

    def find_end(self, start: int, aprox_end: int):
        """
        Encuentra la fila de fin de la tabla basándose en celdas vacías.

        Args:
            start (int): Fila de inicio de la tabla.
            aprox_end (int): Fila aproximada donde termina la búsqueda.

        Returns:
            int or False: Fila de fin si se encuentra, False si no se encuentra.
        """
        end = None 
        for n in range(start, aprox_end):
            if self.sheet_choiced['C' + str(n)].value is None:
                end = n-1
                break
        return end if end!=None else False
    
    def get_start_end_table(self, aprox_start: int, aprox_end: int):
        """
        Obtiene las filas de inicio y fin de la tabla.

        Args:
            aprox_start (int): Fila aproximada donde comienza la búsqueda.
            aprox_end (int): Fila aproximada donde termina la búsqueda.

        Returns:
            list or False: Lista con las filas de inicio y fin si se encuentran, False si no se encuentran.
        """
        end = None
        start = self.find_start(aprox_start, aprox_end)
        if(start):
            end = self.find_end(start, aprox_end)
        return [start, end] if start!=False and end!=None else False